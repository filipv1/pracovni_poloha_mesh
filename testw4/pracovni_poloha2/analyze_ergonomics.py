#!/usr/bin/env python3
"""
Skript pro analýzu CSV dat a vytvoření ergonomického reportu v Excel formátu
Statická analýza úhlů trupu + dynamická analýza frekvence pohybů
"""

import sys
import csv
import argparse
from pathlib import Path
from typing import Dict, List, Tuple
import json

# Excel dependencies
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("Instaluji openpyxl knihovnu...")
    import subprocess
    subprocess.run([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.chart import BarChart, Reference


class ErgonomicAnalyzer:
    """Analyzátor ergonomických dat z CSV souboru"""
    
    def __init__(self, csv_path: str, video_fps: float = 25.0):
        """
        Inicializace analyzátoru
        
        Args:
            csv_path: Cesta k CSV souboru s úhly
            video_fps: FPS videa pro výpočet času
        """
        self.csv_path = Path(csv_path)
        self.video_fps = video_fps
        
        # Kategorie úhlů podle zadání
        self.angle_categories = {
            "-15 a mene (zaklon)": lambda x: x <= -15,
            "-15 az 40 (normal)": lambda x: -15 < x <= 40,
            "40 az 60 (mirny predklon)": lambda x: 40 < x <= 60,
            "60 a vice (vyrazny predklon)": lambda x: x > 60
        }
        
        self.data = []  # List[dict] s frame, angle, time
        self.all_data = []  # Všechna data včetně FALSE hodnot
        self.static_stats = {}
        self.dynamic_stats = {}
        self.missing_periods = []  # Úseky chybějících detekcí
        
    def load_csv_data(self) -> bool:
        """
        Načte data z CSV souboru
        
        Returns:
            True pokud se data načetla úspěšně
        """
        if not self.csv_path.exists():
            print(f"CHYBA: CSV soubor neexistuje: {self.csv_path}")
            return False
            
        print(f"Nacitam data z: {self.csv_path}")
        
        try:
            with open(self.csv_path, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                
                for row in reader:
                    frame_num = int(row['frame'])
                    angle_str = row['úhel_trupu']
                    time_sec = frame_num / self.video_fps
                    
                    # Zpracování hodnot (může být "FALSE" pro nedetekované)
                    if angle_str != "FALSE":
                        angle = float(angle_str)
                        
                        # Platná data pro analýzu
                        self.data.append({
                            'frame': frame_num,
                            'angle': angle,
                            'time': time_sec
                        })
                        
                        # Všechna data
                        self.all_data.append({
                            'frame': frame_num,
                            'angle': angle,
                            'time': time_sec,
                            'detected': True
                        })
                    else:
                        # FALSE hodnoty - chybějící detekce
                        self.all_data.append({
                            'frame': frame_num,
                            'angle': None,
                            'time': time_sec,
                            'detected': False
                        })
            
            print(f"Nacteno {len(self.data)} platnych zaznamu z celkem {len(self.all_data)} framu")
            return True
            
        except Exception as e:
            print(f"CHYBA pri nacitani CSV: {e}")
            return False
    
    def categorize_angle(self, angle: float) -> str:
        """Zařadí úhel do kategorie"""
        for category, condition in self.angle_categories.items():
            if condition(angle):
                return category
        return "Neznámá kategorie"
    
    def calculate_static_analysis(self):
        """Vypočítá statickou analýzu - čas strávený v jednotlivých kategoriích"""
        print("Pocitam statickou analyzu...")
        
        # Inicializace kategorií
        stats = {}
        for category in self.angle_categories.keys():
            stats[category] = {
                'seconds': 0.0,
                'frames': 0
            }
        
        # Počítání času pro každý frame (1/fps sekund na frame)
        frame_duration = 1.0 / self.video_fps
        
        for record in self.data:
            category = self.categorize_angle(record['angle'])
            stats[category]['seconds'] += frame_duration
            stats[category]['frames'] += 1
        
        self.static_stats = stats
        print("Staticka analyza dokoncena")
    
    def detect_transitions(self, target_condition, window_start_sec: float, window_end_sec: float) -> int:
        """
        Detekuje počet vstupů do cílové pozice v daném časovém okně
        
        Args:
            target_condition: Lambda funkce pro test pozice (např. lambda x: x > 60)
            window_start_sec: Začátek okna v sekundách
            window_end_sec: Konec okna v sekundách
            
        Returns:
            Počet vstupů do pozice
        """
        transitions = 0
        was_in_position = None  # None = neznámý stav před oknem
        
        # Najdeme stav před oknem
        for record in self.data:
            if record['time'] < window_start_sec:
                was_in_position = target_condition(record['angle'])
            else:
                break
        
        # Pokud neznáme stav před oknem, začneme s False
        if was_in_position is None:
            was_in_position = False
        
        # Procházíme okno a hledáme transitions
        for record in self.data:
            time_sec = record['time']
            
            # Pouze data v časovém okně
            if window_start_sec <= time_sec <= window_end_sec:
                is_in_position = target_condition(record['angle'])
                
                # Detekce vstupu (transition z False na True)
                if is_in_position and not was_in_position:
                    transitions += 1
                
                was_in_position = is_in_position
            elif time_sec > window_end_sec:
                break  # Už jsme za oknem
        
        return transitions
    
    def find_all_transitions(self, target_condition) -> List[float]:
        """
        Najde všechny přechody do cílové pozice v celém videu
        
        Args:
            target_condition: Lambda funkce pro test pozice
            
        Returns:
            List časů (v sekundách) kdy došlo k přechodu
        """
        transitions = []
        was_in_position = False
        
        for record in self.data:
            is_in_position = target_condition(record['angle'])
            
            # Detekce vstupu (transition z False na True)
            if is_in_position and not was_in_position:
                transitions.append(record['time'])
            
            was_in_position = is_in_position
        
        return transitions
    
    def group_transitions_into_risky_periods(self, transitions: List[float], window_size: float = 60.0) -> int:
        """
        Seskupuje přechody do rizikových období pomocí sliding window přístupu
        
        Args:
            transitions: List časů přechodů
            window_size: Velikost okna v sekundách (default 60.0)
            
        Returns:
            Počet rizikových minut (počet různých minutových úseků s ≥2 přechody)
        """
        if len(transitions) < 2:
            return 0
        
        # Seřadíme přechody podle času
        sorted_transitions = sorted(transitions)
        
        # Najdeme všechny minutové úseky, kde došlo k ≥2 přechodům
        risky_minutes = set()
        
        # Pro každý přechod zkontrolujeme okno následujících 60 sekund
        for i, start_time in enumerate(sorted_transitions):
            end_time = start_time + window_size
            
            # Spočítáme kolik přechodů je v tomto okně (včetně startovního)
            transitions_in_window = 1  # Startovní přechod
            
            for j in range(i + 1, len(sorted_transitions)):
                if sorted_transitions[j] <= end_time:
                    transitions_in_window += 1
                else:
                    break
            
            # Pokud je ≥2 přechodů v okně, označíme tuto minutu jako rizikovou
            if transitions_in_window >= 2:
                # Označíme minutu podle času prvního přechodu
                minute_marker = int(start_time // 60)
                risky_minutes.add(minute_marker)
        
        return len(risky_minutes)
    
    def analyze_missing_periods(self):
        """Analyzuje úseky chybějících detekcí"""
        print("Analyzujem chybejici detekce...")
        
        if not self.all_data:
            return
        
        self.missing_periods = []
        current_start = None
        current_end = None
        
        for record in self.all_data:
            if not record['detected']:
                # Chybějící detekce
                if current_start is None:
                    # Začátek nového úseku
                    current_start = record['time']
                current_end = record['time']  # Neustále aktualizujeme konec
            else:
                # Platná detekce - ukončujeme missing period pokud existoval
                if current_start is not None:
                    self.missing_periods.append({
                        'start_sec': current_start,
                        'end_sec': current_end,
                        'duration_sec': current_end - current_start + (1.0 / self.video_fps)
                    })
                    current_start = None
                    current_end = None
        
        # Pokud video končí missing period
        if current_start is not None:
            self.missing_periods.append({
                'start_sec': current_start,
                'end_sec': current_end,
                'duration_sec': current_end - current_start + (1.0 / self.video_fps)
            })
        
        print(f"Nalezeno {len(self.missing_periods)} useku chybejicich detekci")
    
    def calculate_dynamic_analysis(self):
        """Vypočítá dynamickou analýzu - seskupování rizikových období"""
        print("Pocitam dynamickou analyzu (rizikova obdobi)...")
        
        if not self.data:
            return
        
        total_duration = max(record['time'] for record in self.data)
        print(f"Celkova delka videa: {total_duration:.1f}s")
        
        # Dynamické kategorie s jejich podmínkami
        conditions = {
            "Predklon >60 (>=2/min)": lambda x: x > 60,
            "Predklon >60 (<2/min)": lambda x: x > 60, 
            "Zaklon <=-15 (<2/min)": lambda x: x <= -15
        }
        
        self.dynamic_stats = {}
        
        for category, condition in conditions.items():
            # Najdeme všechny přechody pro tuto kategorii
            transitions = self.find_all_transitions(condition)
            print(f"{category}: nalezeno {len(transitions)} prechodu")
            
            if len(transitions) > 0:
                print(f"  Casy prechodu: {[f'{t:.1f}s' for t in transitions]}")
            
            if category == "Predklon >60 (>=2/min)":
                # Vysoká frekvence - seskupujeme přechody s ≥2 přechody na skupinu
                risky_periods = self.group_transitions_into_risky_periods(transitions, window_size=60.0)
                self.dynamic_stats[category] = risky_periods
                print(f"  Rizikova obdobi: {risky_periods}")
                
            elif category == "Predklon >60 (<2/min)":
                # Nízká frekvence - počítáme období s právě 1 přechodem
                # TODO: Implementovat logiku pro nízkou frekvenci
                self.dynamic_stats[category] = 0
                
            elif category == "Zaklon <=-15 (<2/min)":
                # Nízká frekvence záklonu
                # TODO: Implementovat logiku pro nízkou frekvenci
                self.dynamic_stats[category] = 0
        
        print("Dynamicka analyza dokoncena")
    
    def print_statistics(self):
        """Vytiskne statistiky do konzole"""
        print("\n" + "="*60)
        print("ERGONOMICKA ANALYZA - VYSLEDKY")
        print("="*60)
        
        # Statická analýza
        print("\n1. STATICKA ANALYZA (cas straveny v pozicich):")
        print("-" * 50)
        print(f"{'Kategorie':<30} {'Sekundy':<12} {'Framy':<8}")
        print("-" * 50)
        
        for category, data in self.static_stats.items():
            print(f"{category:<30} {data['seconds']:<12.1f} {data['frames']:<8}")
        
        # Dynamická analýza
        print("\n2. DYNAMICKA ANALYZA (rizikove minuty):")
        print("-" * 50)
        print(f"{'Kategorie':<40} {'Rizikove minuty':<15}")
        print("-" * 50)
        
        for category, risky_minutes in self.dynamic_stats.items():
            print(f"{category:<40} {risky_minutes:<15}")
        
        # Chybějící záznamy
        print("\n3. CHYBEJICI ZAZNAMY:")
        print("-" * 50)
        if self.missing_periods:
            print(f"{'Zacatek (s)':<12} {'Konec (s)':<12} {'Delka (s)':<12}")
            print("-" * 36)
            for period in self.missing_periods:
                print(f"{period['start_sec']:<12.1f} {period['end_sec']:<12.1f} {period['duration_sec']:<12.1f}")
            
            total_missing = sum(p['duration_sec'] for p in self.missing_periods)
            print("-" * 36)
            print(f"Celkem useku: {len(self.missing_periods)}")
            print(f"Celkovy cas chybejicich detekci: {total_missing:.1f}s")
        else:
            print("Zadne chybejici detekce nenalezeny")
    
    def create_excel_report(self, output_path: str):
        """
        Vytvoří Excel report
        
        Args:
            output_path: Cesta k výstupnímu Excel souboru
        """
        print(f"Vytvarem Excel report: {output_path}")
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Ergonomická Analýza"
        
        # Styling
        header_font = Font(bold=True, size=12)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'), 
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        row = 1
        
        # Hlavička reportu
        ws[f'A{row}'] = "ERGONOMICKÁ ANALÝZA TRUPU"
        ws[f'A{row}'].font = Font(bold=True, size=14)
        row += 2
        
        # Základní informace
        ws[f'A{row}'] = "Soubor:"
        ws[f'B{row}'] = str(self.csv_path.name)
        row += 1
        
        ws[f'A{row}'] = "Počet validních záznamů:"
        ws[f'B{row}'] = len(self.data)
        row += 1
        
        if self.data:
            total_time = max(record['time'] for record in self.data)
            ws[f'A{row}'] = "Délka analýzy:"
            ws[f'B{row}'] = f"{total_time:.1f} s"
            row += 2
        
        # 1. STATICKÁ ANALÝZA
        ws[f'A{row}'] = "1. STATICKÁ ANALÝZA"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Hlavičky tabulky
        headers = ['Kategorie úhlu', 'Počet sekund', 'Počet framů']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data statické analýzy
        for category, data in self.static_stats.items():
            ws.cell(row=row, column=1, value=category).border = border
            ws.cell(row=row, column=2, value=round(data['seconds'], 1)).border = border
            ws.cell(row=row, column=3, value=data['frames']).border = border
            row += 1
        
        row += 2
        
        # 2. CHYBĚJÍCÍ ZÁZNAMY
        ws[f'A{row}'] = "2. CHYBĚJÍCÍ ZÁZNAMY (NEDETEKOVANÉ ÚSEKY)"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        if self.missing_periods:
            # Hlavičky tabulky
            headers = ['Začátek (s)', 'Konec (s)', 'Délka trvání (s)']
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=row, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center')
            row += 1
            
            # Data chybějících období
            for period in self.missing_periods:
                ws.cell(row=row, column=1, value=round(period['start_sec'], 1)).border = border
                ws.cell(row=row, column=2, value=round(period['end_sec'], 1)).border = border
                ws.cell(row=row, column=3, value=round(period['duration_sec'], 1)).border = border
                row += 1
            
            # Souhrn
            row += 1
            ws[f'A{row}'] = f"Celkem úseků s chybějící detekcí: {len(self.missing_periods)}"
            ws[f'A{row}'].font = Font(bold=True)
            
            total_missing_time = sum(p['duration_sec'] for p in self.missing_periods)
            ws[f'A{row+1}'] = f"Celkový čas chybějících detekcí: {total_missing_time:.1f} s"
            ws[f'A{row+1}'].font = Font(bold=True)
            row += 3
        else:
            ws[f'A{row}'] = "Žádné chybějící detekce nenalezeny"
            ws[f'A{row}'].font = Font(italic=True, color="00AA00")
            row += 3
        
        # 3. DYNAMICKÁ ANALÝZA
        ws[f'A{row}'] = "3. DYNAMICKÁ ANALÝZA"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        row += 1
        
        # Hlavičky tabulky
        headers = ['Kategorie pohybu', 'Rizikové minuty']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        row += 1
        
        # Data dynamické analýzy
        for category, risky_minutes in self.dynamic_stats.items():
            ws.cell(row=row, column=1, value=category).border = border
            ws.cell(row=row, column=2, value=risky_minutes).border = border
            row += 1
        
        # Automatická šířka sloupců
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Uložení
        wb.save(output_path)
        print(f"Excel report ulozen: {output_path}")


def main():
    """Hlavní funkce"""
    parser = argparse.ArgumentParser(
        description='Analýza CSV dat a vytvoření ergonomického Excel reportu',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Příklady použití:
  %(prog)s data.csv output.xlsx
  %(prog)s data.csv report.xlsx --fps 30
        """
    )
    
    parser.add_argument('csv_file', help='Cesta k CSV souboru s úhly')
    parser.add_argument('excel_output', help='Cesta k výstupnímu Excel souboru')
    parser.add_argument('--fps', type=float, default=25.0, help='FPS videa (default: 25.0)')
    
    args = parser.parse_args()
    
    # Validace
    if not Path(args.csv_file).exists():
        print(f"CHYBA: CSV soubor neexistuje: {args.csv_file}")
        sys.exit(1)
    
    # Vytvoření analyzátoru
    analyzer = ErgonomicAnalyzer(args.csv_file, args.fps)
    
    # Načtení dat
    if not analyzer.load_csv_data():
        sys.exit(1)
    
    # Výpočet analýz
    analyzer.calculate_static_analysis()
    analyzer.analyze_missing_periods()
    analyzer.calculate_dynamic_analysis()
    
    # Zobrazení statistik
    analyzer.print_statistics()
    
    # Vytvoření Excel reportu
    try:
        analyzer.create_excel_report(args.excel_output)
        print(f"\nReport uspesne vytvoren: {args.excel_output}")
        
    except Exception as e:
        print(f"CHYBA pri vytvareni Excel reportu: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()