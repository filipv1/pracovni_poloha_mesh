#!/usr/bin/env python3
"""
Ergonomic Time Analysis
Analyzes CSV output from create_combined_angles_csv_skin.py
Calculates time spent in different angle ranges and sustained positions
Exports results to Excel with organized tables
"""

import pandas as pd
import numpy as np
from pathlib import Path
import sys
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import warnings
warnings.filterwarnings('ignore')

class ErgonomicTimeAnalyzer:
    def __init__(self, csv_file):
        """
        Initialize analyzer
        
        Args:
            csv_file: Input CSV file from create_combined_angles_csv_skin.py
        """
        self.csv_file = Path(csv_file)
        self.fps = None  # Will be read from CSV
        self.df = None
        self.results = {}
        
        # Define angle ranges for different body parts - CZECH LABELS
        self.angle_ranges = {
            'trunk': {
                'extreme_backward': ('≤-15° (silné zakloněni)', lambda x: x <= -15),
                'normal_backward': ('-15° až 40° (normální)', lambda x: -15 < x <= 40),
                'moderate_forward': ('40° až 60° (mírné předkloněni)', lambda x: 40 < x <= 60),
                'extreme_forward': ('≥60° (silné předkloněni)', lambda x: x > 60)
            },
            'neck': {
                'extreme_backward': ('≤-15° (silné zakloněni)', lambda x: x <= -15),
                'normal': ('-15° až 25° (normální)', lambda x: -15 < x <= 25),
                'moderate_forward': ('25° až 60° (mírné předkloněni)', lambda x: 25 < x <= 60),
                'extreme_forward': ('≥60° (silné předkloněni)', lambda x: x > 60)
            },
            'arm': {
                'backward': ('<0° (za tělem)', lambda x: x < 0),
                'normal': ('0° až 40° (normální)', lambda x: 0 <= x <= 40),
                'moderate_elevation': ('40° až 60° (mírné zvednutí)', lambda x: 40 < x <= 60),
                'extreme_elevation': ('≥60° (silné zvednutí)', lambda x: x > 60)
            }
        }
        
        # Czech body part names for display
        self.body_part_czech = {
            'trunk': 'Trup',
            'neck': 'Krk', 
            'left_arm': 'Levá paže',
            'right_arm': 'Pravá paže'
        }
    
    def load_data(self):
        """Load CSV data and validate"""
        if not self.csv_file.exists():
            raise FileNotFoundError(f"CSV file not found: {self.csv_file}")
        
        print(f"Loading CSV: {self.csv_file}")
        self.df = pd.read_csv(self.csv_file)
        
        # Read FPS from CSV (should be same for all rows)
        if 'fps' in self.df.columns:
            self.fps = self.df['fps'].iloc[0]
            print(f"FPS from CSV: {self.fps:.2f}")
        elif 'time_seconds' in self.df.columns and 'frame' in self.df.columns:
            # Calculate FPS from time data
            if len(self.df) > 1:
                time_diff = self.df['time_seconds'].iloc[1] - self.df['time_seconds'].iloc[0]
                if time_diff > 0:
                    self.fps = 1.0 / time_diff
                    print(f"Calculated FPS from time data: {self.fps:.2f}")
                else:
                    self.fps = 30.0
                    print(f"Using default FPS: {self.fps:.2f}")
            else:
                self.fps = 30.0
                print(f"Using default FPS: {self.fps:.2f}")
        else:
            self.fps = 30.0
            print(f"Using default FPS: {self.fps:.2f}")
        
        # Validate required columns
        required_cols = ['frame', 'trunk_angle_skin', 'neck_angle_skin', 
                        'left_arm_angle', 'right_arm_angle']
        missing_cols = [col for col in required_cols if col not in self.df.columns]
        if missing_cols:
            raise ValueError(f"Missing required columns: {missing_cols}")
        
        print(f"Loaded {len(self.df)} frames ({len(self.df)/self.fps:.1f} seconds)")
        print(f"Columns: {list(self.df.columns)}")
        
    def categorize_angles(self):
        """Categorize all angles into ranges"""
        print("\nCategorizing angles into ranges...")
        
        # Add time column if not already present
        if 'time_seconds' not in self.df.columns:
            self.df['time_seconds'] = self.df['frame'] / self.fps
        
        # Categorize trunk angles
        trunk_angles = self.df['trunk_angle_skin']
        self.df['trunk_category'] = trunk_angles.apply(self._categorize_trunk)
        
        # Categorize neck angles  
        neck_angles = self.df['neck_angle_skin']
        self.df['neck_category'] = neck_angles.apply(self._categorize_neck)
        
        # Categorize arm angles
        left_arm_angles = self.df['left_arm_angle']
        right_arm_angles = self.df['right_arm_angle']
        self.df['left_arm_category'] = left_arm_angles.apply(self._categorize_arm)
        self.df['right_arm_category'] = right_arm_angles.apply(self._categorize_arm)
        
        print("Angle categorization completed")
    
    def _categorize_trunk(self, angle):
        """Categorize trunk angle"""
        for category, (label, condition) in self.angle_ranges['trunk'].items():
            if condition(angle):
                return label
        return 'unknown'
    
    def _categorize_neck(self, angle):
        """Categorize neck angle"""
        for category, (label, condition) in self.angle_ranges['neck'].items():
            if condition(angle):
                return label
        return 'unknown'
    
    def _categorize_arm(self, angle):
        """Categorize arm angle"""
        for category, (label, condition) in self.angle_ranges['arm'].items():
            if condition(angle):
                return label
        return 'unknown'
    
    def calculate_basic_times(self):
        """Calculate basic time spent in each angle range"""
        print("Calculating basic time statistics...")
        
        frame_duration = 1.0 / self.fps  # seconds per frame
        
        self.results['basic_times'] = {
            'trunk': self._count_category_time('trunk_category', frame_duration),
            'neck': self._count_category_time('neck_category', frame_duration),
            'left_arm': self._count_category_time('left_arm_category', frame_duration),
            'right_arm': self._count_category_time('right_arm_category', frame_duration)
        }
        
        # Debug: Show angle ranges in data
        print(f"\nDEBUG - Angle ranges in data:")
        print(f"Trunk angles: min={self.df['trunk_angle_skin'].min():.1f}°, max={self.df['trunk_angle_skin'].max():.1f}°")
        print(f"Neck angles: min={self.df['neck_angle_skin'].min():.1f}°, max={self.df['neck_angle_skin'].max():.1f}°")  
        print(f"Left arm angles: min={self.df['left_arm_angle'].min():.1f}°, max={self.df['left_arm_angle'].max():.1f}°")
        print(f"Right arm angles: min={self.df['right_arm_angle'].min():.1f}°, max={self.df['right_arm_angle'].max():.1f}°")
        
        # Print summary
        print("\nBASIC TIME ANALYSIS:")
        for body_part, times in self.results['basic_times'].items():
            print(f"\n{body_part.upper()}:")
            for category, seconds in times.items():
                print(f"  {category}: {seconds:.1f}s")
    
    def _count_category_time(self, column, frame_duration):
        """Count time spent in each category"""
        category_counts = self.df[column].value_counts()
        return {category: count * frame_duration for category, count in category_counts.items()}
    
    def calculate_sustained_positions(self, min_duration=4.0):
        """Calculate time spent in sustained positions (>4 seconds)"""
        print(f"\nCalculating sustained positions (>{min_duration}s)...")
        
        min_frames = int(min_duration * self.fps)
        
        self.results['sustained_times'] = {
            'trunk': self._find_sustained_periods('trunk_category', min_frames),
            'neck': self._find_sustained_periods('neck_category', min_frames),
            'left_arm': self._find_sustained_periods('left_arm_category', min_frames),
            'right_arm': self._find_sustained_periods('right_arm_category', min_frames)
        }
        
        # Print summary
        print(f"\nSUSTAINED POSITION ANALYSIS (>{min_duration}s):")
        for body_part, times in self.results['sustained_times'].items():
            print(f"\n{body_part.upper()}:")
            for category, seconds in times.items():
                print(f"  {category}: {seconds:.1f}s")
    
    def _find_sustained_periods(self, column, min_frames):
        """Find periods where same category lasts for minimum frames"""
        sustained_times = {}
        
        # Get unique categories
        categories = self.df[column].unique()
        
        for category in categories:
            total_sustained_time = 0.0
            current_streak = 0
            
            for _, row in self.df.iterrows():
                if row[column] == category:
                    current_streak += 1
                else:
                    # Streak ended, check if it was long enough
                    if current_streak >= min_frames:
                        total_sustained_time += current_streak / self.fps
                    current_streak = 0
            
            # Check final streak
            if current_streak >= min_frames:
                total_sustained_time += current_streak / self.fps
            
            sustained_times[category] = total_sustained_time
        
        return sustained_times
    
    def export_to_excel(self, output_file="ergonomic_analysis.xlsx"):
        """Export results to Excel with formatted tables"""
        print(f"\nExporting results to Excel: {output_file}")
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Ergonomic Time Analysis"
        
        # Styles
        header_font = Font(bold=True, size=12)
        section_font = Font(bold=True, size=14, color="FFFFFF")
        section_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        border = Border(left=Side(style='thin'), right=Side(style='thin'),
                       top=Side(style='thin'), bottom=Side(style='thin'))
        
        current_row = 1
        
        # Title
        ws.merge_cells(f'A{current_row}:B{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value="ANALÝZA ERGONOMICKÉ ZÁTĚŽE")
        title_cell.font = Font(bold=True, size=16)
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Video info
        total_time = len(self.df) / self.fps
        info_text = f"Video: {self.csv_file.stem} | Snímky: {len(self.df)} | Trvání: {total_time:.1f}s | FPS: {self.fps}"
        ws.merge_cells(f'A{current_row}:B{current_row}')
        info_cell = ws.cell(row=current_row, column=1, value=info_text)
        info_cell.alignment = Alignment(horizontal='center')
        current_row += 3
        
        # 1. Basic time analysis - separate tables for each body part
        current_row = self._add_body_part_tables(ws, current_row, "CELKOVÝ ČAS V JEDNOTLIVÝCH POLOHÁCH", 
                                                self.results['basic_times'], section_font, section_fill, border)
        current_row += 2
        
        # 2. Sustained positions analysis - separate tables for each body part
        current_row = self._add_body_part_tables(ws, current_row, "DLOUHODOBÉ POLOHY (déle než 4 sekundy)", 
                                                self.results['sustained_times'], section_font, section_fill, border)
        
        # Auto-fit columns
        for col_num in range(1, 3):  # We now have 2 columns
            max_length = 0
            column_letter = chr(64 + col_num)  # A, B
            for row in ws.iter_rows(min_col=col_num, max_col=col_num):
                for cell in row:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook
        output_path = Path(output_file)
        wb.save(output_path)
        
        print(f"Excel report saved: {output_path}")
        return output_path
    
    def _add_body_part_tables(self, ws, start_row, main_title, data, section_font, section_fill, border):
        """Add separate tables for each body part"""
        current_row = start_row
        
        # Main section title
        ws.merge_cells(f'A{current_row}:B{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value=main_title)
        title_cell.font = Font(bold=True, size=14, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 2
        
        # Order of body parts
        body_part_order = ['trunk', 'neck', 'left_arm', 'right_arm']
        
        for body_part in body_part_order:
            if body_part in data:
                current_row = self._add_single_body_part_table(
                    ws, current_row, body_part, data[body_part], section_font, section_fill, border
                )
                current_row += 1  # Space between tables
        
        return current_row
    
    def _add_single_body_part_table(self, ws, start_row, body_part, times, section_font, section_fill, border):
        """Add a table for a single body part"""
        current_row = start_row
        
        # Body part title
        body_part_display = self.body_part_czech.get(body_part, body_part.replace('_', ' ').title())
        ws.merge_cells(f'A{current_row}:B{current_row}')
        subtitle_cell = ws.cell(row=current_row, column=1, value=body_part_display)
        subtitle_cell.font = Font(bold=True, size=12)
        subtitle_cell.fill = PatternFill(start_color="D9E2F3", end_color="D9E2F3", fill_type="solid")
        subtitle_cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Headers (only 2 columns)
        headers = ['Rozsah polohy', 'Čas (sekundy)']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Data rows
        for position, seconds in times.items():
            # Position
            ws.cell(row=current_row, column=1, value=position).border = border
            # Time
            time_cell = ws.cell(row=current_row, column=2, value=f"{seconds:.1f}")
            time_cell.border = border
            time_cell.alignment = Alignment(horizontal='right')
            
            current_row += 1
        
        return current_row
    
    def _add_time_table(self, ws, start_row, title, data, section_font, section_fill, border):
        """Add a time analysis table to worksheet"""
        current_row = start_row
        
        # Section title
        ws.merge_cells(f'A{current_row}:D{current_row}')
        title_cell = ws.cell(row=current_row, column=1, value=title)
        title_cell.font = section_font
        title_cell.fill = section_fill
        title_cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Headers
        headers = ['Část těla', 'Rozsah polohy', 'Čas (sekundy)', 'Procenta']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = Font(bold=True)
            cell.border = border
            cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Calculate total time for percentages
        total_time = len(self.df) / self.fps
        
        # Data rows
        for body_part, times in data.items():
            # Use Czech body part name
            body_part_display = self.body_part_czech.get(body_part, body_part.replace('_', ' ').title())
            
            for position, seconds in times.items():
                percentage = (seconds / total_time) * 100
                
                # Body part
                ws.cell(row=current_row, column=1, value=body_part_display).border = border
                # Position
                ws.cell(row=current_row, column=2, value=position).border = border
                # Time
                time_cell = ws.cell(row=current_row, column=3, value=f"{seconds:.1f}")
                time_cell.border = border
                time_cell.alignment = Alignment(horizontal='right')
                # Percentage
                pct_cell = ws.cell(row=current_row, column=4, value=f"{percentage:.1f}%")
                pct_cell.border = border
                pct_cell.alignment = Alignment(horizontal='right')
                
                current_row += 1
        
        return current_row
    
    def run_analysis(self, output_excel="ergonomicka_analyza.xlsx"):
        """Run complete analysis"""
        print("ANALÝZA ERGONOMICKÉ ZÁTĚŽE")
        print("=" * 50)
        
        try:
            self.load_data()
            self.categorize_angles()
            self.calculate_basic_times()
            self.calculate_sustained_positions(min_duration=4.0)
            excel_path = self.export_to_excel(output_excel)
            
            print(f"\n✅ ANALÝZA ÚSPĚŠNĚ DOKONČENA")
            print(f"📊 Excel report: {excel_path}")
            print(f"📈 Celková délka videa: {len(self.df)/self.fps:.1f} sekund")
            print(f"🎬 FPS videa: {self.fps:.2f}")
            
            return excel_path
            
        except Exception as e:
            print(f"❌ CHYBA: {e}")
            return None

def main():
    """Main execution function"""
    # Default input file
    default_csv = "combined_angles_skin_5614.csv"
    
    if len(sys.argv) > 1:
        input_csv = sys.argv[1]
    else:
        input_csv = default_csv
    
    # Check if input file exists
    if not Path(input_csv).exists():
        print(f"ERROR: Input CSV file not found: {input_csv}")
        print("\nAvailable CSV files:")
        csv_files = list(Path(".").glob("*angles*.csv"))
        if csv_files:
            for csv_file in csv_files:
                print(f"  - {csv_file}")
            print(f"\nUsage: python {sys.argv[0]} <csv_file>")
        else:
            print("  No CSV files found in current directory")
            print("  Run create_combined_angles_csv_skin.py first")
        return
    
    print(f"Input CSV: {input_csv}")
    
    # Create output filename
    input_path = Path(input_csv)
    output_excel = f"ergonomicka_analyza_{input_path.stem}.xlsx"
    
    # Run analysis
    analyzer = ErgonomicTimeAnalyzer(input_csv)
    result = analyzer.run_analysis(output_excel)
    
    if result:
        print(f"\n🎯 PŘIPRAVENO K PROHLÉDNUTÍ:")
        print(f"Otevřete {result} pro detailní ergonomickou analýzu")
        print(f"\nAnalýza obsahuje:")
        print(f"• Čas strávený v jednotlivých úhlových rozsazích")
        print(f"• Analýzu dlouhodobých poloh (>4 sekundy)")
        print(f"• Procentuální rozdělení")
        print(f"• Přehledné tabulky pro snadnou interpretaci")

if __name__ == "__main__":
    main()